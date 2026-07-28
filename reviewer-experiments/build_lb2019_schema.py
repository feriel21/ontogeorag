#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_lb2019_schema.py  (ADDITIF — ne touche pas run11)
Derive le schema LB2019 depuis Table_Supplementary_1_V2.xlsx (reproductible).
Remplace LB2019_DESCRIPTORS / LB2019_REFERENCE_EDGES ecrits a la main.
Sort: lb2019_schema_generated.json + lb2019_descriptors.json + lb2019_reference_edges.json
Usage: python build_lb2019_schema.py --xlsx Table_Supplementary_1_V2.xlsx --outdir ./schema_generated
"""
import argparse, json, hashlib, re, sys, datetime
from pathlib import Path
try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl --break-system-packages")

EDGE_SEP = " - "

def file_sha256(p: Path) -> str:
    h = hashlib.sha256(); h.update(p.read_bytes()); return h.hexdigest()[:16]

def parse_refs(val):
    if val is None: return []
    return [int(x) for x in re.findall(r"\d+", str(val))]

def split_edge_label(label: str):
    parts = label.split(EDGE_SEP)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip(), False
    return parts[0].strip(), EDGE_SEP.join(parts[1:]).strip(), True

def load_descriptor_property_map(wb):
    ws = wb["Edges_OrganizedByDescriptor"]
    desc2prop, current_prop = {}, None
    for r in ws.iter_rows(min_row=1, values_only=True):
        a, b, c, d = r[0], r[1], r[2], r[3]
        if a: current_prop = str(a).strip()
        if b and (c is None) and (d is None) and current_prop:
            name = str(b).strip()
            if name.lower() != "descriptor":
                desc2prop[name] = current_prop
    return desc2prop

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--outdir", default="./schema_generated")
    args = ap.parse_args()
    xlsx = Path(args.xlsx)
    if not xlsx.exists(): sys.exit(f"introuvable : {xlsx}")
    outdir = Path(args.outdir); outdir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    desc2prop = load_descriptor_property_map(wb)
    ws = wb["EDGES_TABLE"]
    edges, flagged = [], []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        label = r[0]
        if not label: continue
        src, tgt, flag = split_edge_label(str(label))
        e = {"source": src, "target": tgt,
             "type": (str(r[1]).strip() if r[1] else None),
             "refs": parse_refs(r[2]),
             "comment": (str(r[3]).strip() if r[3] else None),
             "xlsx_row": i, "raw_label": str(label).strip()}
        edges.append(e)
        if flag: flagged.append(e)
    nodes = {}
    for e in edges:
        for n in (e["source"], e["target"]):
            if n not in nodes:
                nodes[n] = {"label": n,
                            "role": "descriptor" if n in desc2prop else "process_or_control",
                            "property": desc2prop.get(n)}
    descriptors = sorted([n for n, v in nodes.items() if v["role"] == "descriptor"])
    processes   = sorted([n for n, v in nodes.items() if v["role"] == "process_or_control"])
    schema = {"_meta": {
            "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "source_file": xlsx.name, "source_sha256_16": file_sha256(xlsx),
            "n_edges": len(edges), "n_nodes": len(nodes),
            "n_descriptors": len(descriptors), "n_processes_or_controls": len(processes),
            "n_directed": sum(e["type"] == "Directed" for e in edges),
            "n_undirected": sum(e["type"] == "Undirected" for e in edges),
            "n_flagged_for_manual_review": len(flagged),
            "note": "Types Directed/Undirected uniquement. Relations typees pas d'ici."},
        "nodes": nodes, "edges": edges, "flagged_edges": flagged}
    (outdir / "lb2019_schema_generated.json").write_text(
        json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8")
    (outdir / "lb2019_descriptors.json").write_text(
        json.dumps({"descriptors": descriptors, "descriptor_to_property": desc2prop},
                   ensure_ascii=False, indent=2), encoding="utf-8")
    (outdir / "lb2019_reference_edges.json").write_text(
        json.dumps([{"source": e["source"], "target": e["target"], "type": e["type"],
                     "refs": e["refs"], "xlsx_row": e["xlsx_row"]} for e in edges],
                   ensure_ascii=False, indent=2), encoding="utf-8")
    m = schema["_meta"]
    print("=== LB2019 schema derive (reproductible) ===")
    print(f"  edges  : {m['n_edges']} (Directed {m['n_directed']}/Undirected {m['n_undirected']})")
    print(f"  noeuds : {m['n_nodes']} (descripteurs {m['n_descriptors']}/processus {m['n_processes_or_controls']})")
    print(f"  a revoir main (split != 2) : {m['n_flagged_for_manual_review']}")
    for e in flagged: print(f"     row {e['xlsx_row']}: {e['raw_label']}")
    print(f"  -> {outdir}/  (3 fichiers JSON)")

if __name__ == "__main__":
    main()
